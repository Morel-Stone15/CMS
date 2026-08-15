import io
import openpyxl
from openpyxl import Workbook
from werkzeug.security import generate_password_hash
from models import db, Member
from services.qr_service import generate_qr_code, generate_pin
from services.email_service import log_action

def generate_members_excel():
    """Generate Excel sheet of all members."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Membres CLUB TECH"
    
    headers = [
        "N° Membre", "Nom", "Prénom", "Filière", 
        "Niveau", "Email", "Téléphone", "Statut", "Rôle Bureau", "Date Inscription"
    ]
    ws.append(headers)
    
    members = Member.query.all()
    for m in members:
        ws.append([
            m.member_number,
            m.last_name,
            m.first_name,
            m.major,
            m.level,
            m.email,
            m.phone,
            m.status or 'actif',
            "Oui" if m.is_bureau else "Non",
            m.created_at.strftime("%Y-%m-%d %H:%M:%S") if m.created_at else ""
        ])
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def import_members_from_excel(file_stream, operator='Bureau'):
    """Import members from Excel file."""
    wb = openpyxl.load_workbook(file_stream)
    ws = wb.active
    
    imported_count = 0
    errors = []
    
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[2] or not row[6]: # Needs last name and email
            continue
            
        if row[1] is not None:
            raw_pin = str(row[1]).split('.')[0].strip()
            pin = raw_pin if raw_pin else generate_pin()
        else:
            pin = generate_pin()
        last_name = row[2]
        first_name = row[3]
        major = row[4] or "Général"
        level = row[5] or "N/A"
        email = row[6]
        phone = str(row[7]) if row[7] is not None else ""
        
        if Member.query.filter_by(email=email).first():
            errors.append(f"Ligne {r_idx}: Email '{email}' déjà existant, sauté.")
            continue
            
        if not member_number:
            from blueprints.auth import generate_member_number
            member_number = generate_member_number()
            
        qr_path = generate_qr_code(member_number)
        
        new_m = Member(
            member_number=member_number,
            pin=generate_password_hash(pin[:6]),
            last_name=last_name,
            first_name=first_name,
            major=major,
            level=level,
            email=email,
            phone=phone,
            qr_code_path=qr_path,
            is_bureau=False
        )
        db.session.add(new_m)
        db.session.commit()
        imported_count += 1
        
    log_action(operator, f"Importation réussie de {imported_count} membres depuis un fichier Excel.")
    return imported_count, errors
